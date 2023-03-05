# -*- coding: utf8 -*-
import re
from openpyxl import load_workbook


class DataValue:
    def __init__(self, name, number=0, count=1):
        self.name = name
        self.values = []
        self.sections = {}
        self.count = count
        self.number = number


class Sections:
    not_check_result = {}
    check_result = {}



class Counter:
    def __init__(self):
        self.counter = 0
        self.wrong_name = []
        self.wrong_num = []

    def __call__(self, *args, **kwargs):
        if args[0] not in self.wrong_num:
            self.counter += 1
            self.wrong_num.append(args[0])
            self.wrong_name.append(args[1])

    def withdraw(self):
        for i in range(len(self.wrong_num)):
            print(f"{self.wrong_num[i]}.{self.wrong_name[i]}")


# class DataVals:

t_column = {1: "A",  # № п/п
            2: "B",  # Обоснование
            3: "C",  # Наименование работ и затрат
            4: "F",  # Единица измерения
            5: "G",  # Количество: на единицу
            6: "H",  # Количество: коэффициенты
            7: "I",  # Количесвто: всего с учетом коэффициентов
            8: "J",  # Сметная стоимость в базисном уровне цен, руб.: на единицу
            9: "K",  # Сметная стоимость в базисном уровне цен, руб.: коэффициенты
            10: "L",  # Сметная стоимость в базисном уровне цен, руб.: всего
            11: "M",  # Индексы
            12: "N"  # Сметная стоимость в текущем уровне цен, руб.
            }

expressions_ter = [

    {"regular": r"^ОТ$|^ЭМ$|^в т.ч. ОТм$|^М$",
     "columns": [[t_column[8], t_column[9], t_column[10]], [t_column[10], t_column[11], t_column[12]]],
     "function": 1
     },
    {"regular": r"^ЗТ$|^ЗТм$",
     "columns": [[t_column[5], t_column[6], t_column[7]]],
     "function": 1
     },
    {"regular": r"^Итого\s+по\s+расценке$",
     "columns": [t_column[8], t_column[10], t_column[12]],
     "strings": r"^ОТ$|^ЭМ$|^М$",
     "function": 2
     },
    {"regular": r"^ФОТ$",
     "columns": [t_column[10], t_column[12]],
     "strings": r"^ОТ$|^в т.ч. ОТм$",
     "function": 2
     },
    {"regular": r"^.{18,}$",
     "columns": [[t_column[5], t_column[6], t_column[7]]],
     "function": 1
     },
    {"regular": r"^Всего\s+по\s+позиции$",
     "columns": [t_column[10], t_column[12]],
     "strings": r"^Итого по расценке$|^.{18,}$",
     "function": 2
     }
]

expressions_tssc = [

    {"regular": r"^.{18,}$",
     "columns": [[t_column[8], t_column[9], t_column[10]], [t_column[10], t_column[11], t_column[12]]],
     "function": 1
     },
    {"regular": r"^Всего\s+по\s+позиции$",
     "columns": [t_column[10], t_column[12]],
     "strings": r"\b(?:(?!Всего|по|позиции)\w)+\b",
     "function": 2
     }
]

counter_wrong = Counter()


def main():
    num_str = 0
    total_num_str = 5000  # количество строк в таблице
    wb = load_workbook('15_02-05-03_Вентиляция.xlsx')
    ws = wb.active
    sections = Sections()

    while num_str < total_num_str:
        num_str += 1
        justification = str(ws[f"{t_column[2]}{num_str}"].value)
        if re.compile(r"^ТЕР.+$").findall(justification):
            ter = DataValue(str(ws[f"{t_column[2]}{num_str}"].value),
                            ws[f"{t_column[1]}{num_str}"].value,
                            ws[f"{t_column[7]}{num_str}"].value)  # ищет название Теров
            print(f"{ter.number}.{ter.name}")
            num_str = get_rationale(num_str + 1, ws, ter, r"\bВсего\s+по\s+позиции\b")
            check_ter_tssc_tc(ter, expressions_ter)
            sections.not_check_result[ter.number] = ter.values
        elif re.compile(r"^ТЦ.+$|^ТССЦ.+$|^Прайс-лист$").findall(justification):
            tssc = DataValue(str(ws[f"{t_column[2]}{num_str}"].value),
                             ws[f"{t_column[1]}{num_str}"].value,
                             ws[f"{t_column[7]}{num_str}"].value)  # ищет название Теров
            print(f"{tssc.number}.{tssc.name}")
            num_str = get_rationale(num_str, ws, tssc, r"\bВсего\s+по\s+позиции\b")
            check_ter_tssc_tc(tssc, expressions_tssc)
            sections.not_check_result[tssc.number] = tssc.values


    print(f"Количество неправильных: {counter_wrong.counter}")
    counter_wrong.withdraw()


def check_ter_tssc_tc(ter, expressions):
    for position in ter.values:
        for exp in expressions:
            if re.compile(exp["regular"]).findall(position["C"]):
                if exp["function"] == 1:
                    result = check_values_string(position, ter.count, exp["columns"])
                    print_result(result, ter, position["C"])
                elif exp["function"] == 2:
                    result = checkSumColumn(position, ter, exp["columns"], exp["strings"])
                    print_result(result, ter, position["C"])


def print_result(val, ter, pos):
    if val:
        string_check = "OK!"
    elif val is False:
        string_check = "Неправильно!"
        counter_wrong(ter.number, ter.name)
    else:
        string_check = "Не правильные данные!"
        counter_wrong(ter.number, ter.name)

    print(f"\t{pos}...{string_check}")


def get_rationale(num_str, ws, ter, end_flag):
    """Теры, тссц и тц"""
    while True:
        d_values = get_values(ws, num_str)
        if d_values:
            ter.values.append(d_values)
        if re.compile(end_flag).findall(str(ws[f"{t_column[3]}{num_str}"].value)):
            return num_str
        num_str += 1



def checkSumColumn(position, ter, columns, strings):
    """Проверяет  сумму в колонках"""

    for col in columns:
        sum_values = 0
        for val in ter.values:
            if re.compile(strings).findall(val["C"]):
                sum_values += val[col]
        try:
            if round(sum_values, 2) != position[col]:
                return False
        except TypeError:
            return None
    return True


def check_values_string(values, count, columns):
    """Проверяет данные построчно"""

    if "F" in values and values["F"] == "%":
        count = 1

    for col1, col2, col3 in columns:
        if col2 not in values:
            values[col2] = 1
        try:
            if int(round(float(count) * values[col1] * values[col2], 2)) != int(values[col3]):
                return False
            count = 1
        except TypeError:
            return None
    return True


def get_values(ws, num_str):
    """Получает значения в строчке, если есть, и добавляет их в словарь"""
    d_values = {}
    for i in t_column.values():
        if ws[f"{i}{num_str}"].value:
            d_values[i] = ws[f"{i}{num_str}"].value
    if checkisfull(d_values) is False:
        return False
    return d_values


def checkisfull(d_values):
    """Проверка если ли числовые значения и их минуму два"""
    return len([val for val in d_values.values() if isinstance(val, (float, int))]) >= 1


if __name__ == "__main__":
    main()
